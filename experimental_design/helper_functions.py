import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from scipy.optimize import curve_fit

from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import RBF
from scipy.optimize import minimize, curve_fit
from scipy.stats import norm 
import time
from IPython.display import clear_output
import random
from openpyxl import load_workbook

def req():
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    from scipy.optimize import curve_fit

    from sklearn.gaussian_process import GaussianProcessRegressor
    from sklearn.gaussian_process.kernels import RBF
    from scipy.optimize import minimize
    from scipy.stats import norm 
    
    

def measure_some_system(x1, x2):
    y = ((x1**2-1)*(x2**2-5)+x1**2 +x2**2-6)/((x1-0.1)**2 +x2**2 +2)**2
    return y

def read_excel_and_measure(sheet, delay = 7):
    df = pd.read_excel('LabBook.xlsx', sheet_name = sheet)
    x1 = df.iloc[:,0]
    x2 = df.iloc[:,1]
    y = measure_some_system(x1, x2)

    workbook = load_workbook(filename='LabBook.xlsx')
    sheet = workbook[sheet] 

    # Specify the column and data to write
    column_letter_y = "C"  
    # Write data to the specified column, starting from row 2
    for row, value in enumerate(y, start=2):  # start=2 assumes headers are in the first row
        sheet[f"{column_letter_y}{row}"] = value

    # Save the workbook
    workbook.save(filename='LabBook.xlsx')


    print("YOUR MEASUREMENT IS BEING PROCESSED. PLEASE WAIT! ")
    time.sleep(0.5)
    ascii_art_1 = ["YOUR MEASUREMENT IS BEING PROCESSED. PLEASE WAIT! ",
                 " ",                   
        "                    _..._                  ",   
        "                   //''\\\                 ", 
        "                   ||. .||                 ",  
        "                   |\ _ /|      (          ",  
        "                  .-/\ /\-.     )   |      ",   
        "                 |  ` \ '  |    _   |      ",   
        "           (     | |  |  | |    H=--+-     ",   
        "           ))    | |__|[ | |    U   |      ",   
        "           __    \___(_3/ /     )   |      ",   
        "-|_H_H_|---||---------|!|/------|---|---.  ",
        " |_U_U_|  /__\        |_|      _[_ _|__  \ ",
        "------------------------------------------`"]
    
    ascii_art_2 = ["YOUR MEASUREMENT IS BEING PROCESSED. PLEASE WAIT! ",
                 " ",                   
        "                    _..._                  ",   
        "                   //''\\\                 ", 
        "                   ||o o||                 ",  
        "                   |\ _ /|      )          ",  
        "                  .-/\ /\-.     (   |      ",   
        "                 |  ` \ '  |    _   |      ",   
        "           )     | |  |  | |    H=--+-     ",   
        "           ((    | |__|[ | |    U   |      ",   
        "           __    \___(_3/ /     (   |      ",   
        "-|_H_H_|---||---------|!|/------|---|---.  ",
        " |_U_U_|  /__\        |_|      _[_ _|__  \ ",
        "------------------------------------------`"]
    #display_ascii_art_slowly(ascii_art, delay = delay)
    iterate_ascii_art([ascii_art_1, ascii_art_2], t = delay)


    print('YOUR MEASUREMENTS ARE FINISHED!')



def lab_measurement(x1, x2, delay = 7):
    y = measure_some_system(x1, x2)
    print("YOUR MEASUREMENT IS BEING PROCESSED. PLEASE WAIT! ")
    time.sleep(0.5)
    ascii_art_1 = ["YOUR MEASUREMENT IS BEING PROCESSED. PLEASE WAIT! ",
                 " ",                   
        "                    _..._                  ",   
        "                   //''\\\                 ", 
        "                   ||. .||                 ",  
        "                   |\ _ /|      (          ",  
        "                  .-/\ /\-.     )   |      ",   
        "                 |  ` \ '  |    _   |      ",   
        "           (     | |  |  | |    H=--+-     ",   
        "           ))    | |__|[ | |    U   |      ",   
        "           __    \___(_3/ /     )   |      ",   
        "-|_H_H_|---||---------|!|/------|---|---.  ",
        " |_U_U_|  /__\        |_|      _[_ _|__  \ ",
        "------------------------------------------`"]
    
    ascii_art_2 = ["YOUR MEASUREMENT IS BEING PROCESSED. PLEASE WAIT! ",
                 " ",                   
        "                    _..._                  ",   
        "                   //''\\\                 ", 
        "                   ||o o||                 ",  
        "                   |\ _ /|      )          ",  
        "                  .-/\ /\-.     (   |      ",   
        "                 |  ` \ '  |    _   |      ",   
        "           )     | |  |  | |    H=--+-     ",   
        "           ((    | |__|[ | |    U   |      ",   
        "           __    \___(_3/ /     (   |      ",   
        "-|_H_H_|---||---------|!|/------|---|---.  ",
        " |_U_U_|  /__\        |_|      _[_ _|__  \ ",
        "------------------------------------------`"]
    #display_ascii_art_slowly(ascii_art, delay = delay)
    iterate_ascii_art([ascii_art_1, ascii_art_2], t = delay)

    print("")
    print("YOUR MEASUREMENT IS FINISHED!")
    print("The result is: ")
    print(y)

    return y
def linear(X, a, b, c):
    return X[0]*a + X[1] * b + c

def quadratic(X, a, b, c, d, e, f):
    return a * X[0]**2 + b * X[1]**2 + c * X[0] * X[1] + d * X[0] + e * X[1] + f

def gaussian(X, A, mu_1, sigma_1, mu_2, sigma_2):
    exponent = -((X[0] - mu_1)**2 / (2 * sigma_1**2)) - ((X[1] - mu_2)**2 / (2 * sigma_2**2))
    return A * np.exp(exponent)


def function_fit(model_function, sheet):
    df = pd.read_excel('LabBook.xlsx', sheet_name= sheet)
    print(df)
    if model_function =='linear':
        f = linear
        popt = [1,1,1]
    elif model_function =='quadratic':
        f = quadratic
        popt = [1,1,1,1,1,1]
    elif model_function =='gaussian':
        f = gaussian
        popt = [1,1,1,1,1]



    popt, pcov = curve_fit(f, [df.iloc[:, 0], df.iloc[:,1]], df.iloc[:,2], p0 = popt)

    return popt, pcov, df.iloc[:, 0], df.iloc[:,1]

def plot_gpr(gpr, x1, x2, i, resolution = 100, bounds = [-3,3], minimize =False):
    # gpr: gaussian process regressor object
    # x1 and x2: the points gpr was fitted on
    # i: iteration number
    resolution = 100
    x1_ =np.linspace(bounds[0],bounds[1], resolution)
    x2_ = x1_
    X1_, X2_ = np.meshgrid(x1_, x2_)
    X_all = [[X1_[i,j], X2_[i,j]] for i in range(resolution) for j in range(resolution)]
    Y_all = gpr.predict(X_all, return_std = False)
    Y_all = Y_all.reshape((resolution,resolution))

    plt.figure(figsize=(5,5))
    c = plt.pcolormesh(X1_, X2_, Y_all, cmap = 'twilight')
    cbar = plt.colorbar(c)
    cbar.ax.set_ylabel('Output-Effect [au*au]', rotation=-90, va="bottom")
    #ax1.contour(x1_system, x2_system, y_system)
    plt.xlabel('Factor X1 [au]')
    plt.ylabel('Factor X2 [au]')
    plt.title(f'Current Knowledge after {i}-th measurement ')
    for i in range(len(x1)):
        plt.plot(x1[i], x2[i], 'x', color = 'black')

    if minimize:
        y = gpr.predict(X_all)
        idx = np.where(y == min(y))[0][0]
        minx1 = X_all[idx][0]
        minx2 = X_all[idx][1]   
        miny = min(y)
        plt.plot(minx1, minx2, 'x', markersize = 15, color = 'g')
        plt.savefig(f'ResultsGraphs\Bayesian_Minimum.png')
        plt.show()
        return minx1, minx2, miny
    else:
        plt.savefig(f'ResultsGraphs\Bayesian-{i:02d}')
        plt.show()

def show_parameter_space(): 
    print("THIS IS YOUR PARAMETER SPACE : ")
    X, Y, = np.meshgrid(np.zeros(5), np.zeros(5))
    Z = Y
    plt.figure(figsize =(5,5))
    c = plt.pcolormesh(X,Y,Z, cmap = 'twilight')#, vmin = 0, vmax = 0)
    cbar = plt.colorbar(c)
    plt.xlim(-3,3)
    plt.ylim(-3,3)
    plt.xlabel('Factor X1')
    plt.ylabel('Factor X2')
    cbar.ax.set_ylabel('Output-Effect [au*au]', rotation=-90, va="bottom")
    plt.title(f'ParameterSpace')
    plt.grid()
    plt.show()


    

    

def bayesian_optimizer(sheet= 'Bayesian Optimization', output =True):
    df = pd.read_excel('LabBook.xlsx', sheet_name= sheet)
    df = df.dropna(how='all')
    print("YOUR DATA :")
    print(df)
    x1 = df.iloc[:,0]
    x2 = df.iloc[:,1]
    y = df.iloc[:,2]

    # define kernel and regressor
    kernel = RBF(length_scale = 1)
    gpr = GaussianProcessRegressor(kernel=kernel,
            random_state=0)
    

    X = [[x1[i], x2[i]] for i in range(len(x1))]
    gpr.fit(X,y)
    #print(output)
    if output:
        # plot the current output
        i = len(x1)
        plot_gpr(gpr, x1, x2, i)

        next_sample_1 = minimize(lambda x: upper_confidence_bound(x.reshape(1, -1), X, y, gpr, kappa = 0.5),
                           x0=np.random.uniform(0, 1, size=(2,)),
                           bounds=[(-3, 3), (-3, 3)],
                           method='L-BFGS-B').x
        
        next_sample_2 = minimize(lambda x: upper_confidence_bound(x.reshape(1, -1), X, y, gpr, kappa = 1.5),
                           x0=np.random.uniform(0, 1, size=(2,)),
                           bounds=[(-3, 3), (-3, 3)],
                           method='L-BFGS-B').x #np.random.uniform(-3, 3, size=(2,))#

        print(' YOUR NEXT MEASUREMENT POINTS SHOULD BE: ')
        print(' Measurement 1: ')
        print(f' Factor 1 : {next_sample_1[0]}')
        print(f' Factor 2 : {next_sample_1[1]}')
        print(' Measurement 2: ')
        print(f' Factor 1 : {next_sample_2[0]}')
        print(f' Factor 2 : {next_sample_2[1]}')
        print('WRITE DOWN THE POINTS IN THE LABBOOK, ')
        print('MEASURE THEM AND ')
        print('DONT FORGET TO CLOSE THE EXCEL FILE AGAIN.')
    
    else: 
        return gpr, x1, x2
        


def minimize_and_visualize(opt, cov, model_function, type, x1original=0, x2original=0 ):
    if model_function =='linear':
        f = linear
        traditional = True
    elif model_function =='quadratic':
        f = quadratic
        traditional = True
    elif model_function =='gaussian':
        f = gaussian
        traditional = True
    elif model_function =='gpr':
        traditional = False

    if traditional:
        x1 = np.linspace(-3,3,50)
        x2 = np.linspace(-3,3,50)
        X1, X2 = np.meshgrid(x1, x2)
        X1_all = X1.reshape(-1, 1)
        X2_all = X2.reshape(-1,1)

        y = f([X1, X2], *opt)

        # minimize
        y_ = y.reshape(-1,1)
        idx = np.where(y_ == min(y_))[0][0]
        minx1 = X1_all[idx]
        minx2 = X2_all[idx]
        miny = min(y_)

        # visualize
        plt.figure(figsize=(5,5))
        c = plt.pcolormesh(X1, X2, y, cmap = 'twilight')
        cbar = plt.colorbar(c)
        plt.plot(minx1, minx2, 'x', markersize = 15, color = 'g')
        plt.plot(x1original, x2original, 'x', color = 'black')
        plt.xlabel('Factor X1')
        plt.ylabel('Factor X2')

        cbar.ax.set_ylabel('Output-Effect [au*au]', rotation=-90, va="bottom")
        plt.title(f'Results {type} Approach')
        plt.savefig(f'ResultsGraphs\{type}.png')
        plt.show()
        
    else:
        gpr, x1, x2 = bayesian_optimizer(sheet = opt, output = False)
        minx1, minx2, miny = plot_gpr(gpr, x1, x2, 20, minimize =True)

    print('Your minimum can be found at: ')
    print(f'Factor 1: {minx1}')
    print(f'Factor 2: {minx2}')
    print(f'It has the value: {miny}')




def gimme_random_experiments(no_of_experiments, sheet):
    bound_1 = -3
    bound_2 = 3

    x1 = [random.uniform(bound_1, bound_2) for _ in range(no_of_experiments)]
    x2 = [random.uniform(bound_1, bound_2) for _ in range(no_of_experiments)]



    # Load the workbook and select the active worksheet
    file_path = "LabBook.xlsx"
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet] 

    # Specify the column and data to write
    column_letter_x1 = "A"  
    column_letter_x2 = 'B'
    # Write data to the specified column, starting from row 2
    for row, value in enumerate(x1, start=2):  # start=2 assumes headers are in the first row
        sheet[f"{column_letter_x1}{row}"] = value
    for row, value in enumerate(x2, start=2):  # start=2 assumes headers are in the first row
        sheet[f"{column_letter_x2}{row}"] = value    

    # Save the workbook
    workbook.save(filename=file_path)


def iterate_ascii_art(arts, t=10):
    progress_display = ['|', '/', '-', '\\',]
    for i in range(t):
        pi = progress_display[i%len(progress_display)]
        clear_output(wait=True)
        a = arts[0].copy()
        a[0]+=pi
        print("\n".join(a))
        time.sleep(0.5)

        clear_output(wait=True)
        a = arts[1].copy()
        a[0]+=pi
        print("\n".join(a))
        time.sleep(0.5)


# Function to display ASCII art line by line
def display_ascii_art_slowly(art, delay=0.5):
    for i in range(1, len(art) + 1):
        clear_output(wait=True)  # Clear previous output
        #print("\n".join(art[:i]))  # Print up to the i-th line
        time.sleep(delay)  # Wait before printing the next line



def plot_the_system():
    X1 = np.linspace(-3, 3, 2000)
    X2 = np.linspace(-3, 3, 2000)

    x1_system, x2_system = np.meshgrid(X1, X2)
    y_system = measure_some_system(x1_system,x2_system)


    plt.figure(figsize=(10,5))
    ax1 = plt.subplot(121)
    c = ax1.pcolormesh(x1_system, x2_system, y_system, cmap = 'twilight')
    cbar = plt.colorbar(c)
    cbar.ax.set_ylabel('Output-Effect [au*au]', rotation=-90, va="bottom")
    ax1.contour(x1_system, x2_system, y_system)
    ax1.set_xlabel('Factor X1 [au]')
    ax1.set_ylabel('Factor X2 [au]')
    ax1.set_title('Exemplary System - Contour')

    ax2 = plt.subplot(122, projection = '3d')
    ax2.plot_surface(x1_system, x2_system, y_system, cmap = 'twilight')
    ax2.set_xlabel('Factor X1 [au]')
    ax2.set_ylabel('Factor X2 [au]')
    ax2.set_zlabel('Output-Effect [au*au]')
    ax2.set_title('Exemplary System - 3D surface')

    return ax1, ax2


# Acquisition function: Expected Improvement
def expected_improvement(X, X_sample, y_sample, gp, xi=0.01):
    X = X.reshape(-1, 2)  # Ensure correct shape for input points
    mu, sigma = gp.predict(X, return_std=True)
    mu_sample = gp.predict(X_sample)

    # Calculate the improvement
    sigma = sigma.reshape(-1, 1)
    with np.errstate(divide='warn'):
        imp = mu - np.min(mu_sample) - xi
        Z = imp / sigma
        ei = imp * norm.cdf(Z) + sigma * norm.pdf(Z)
        ei[sigma == 0.0] = 0.0
    return ei


# Aquisition function: Upper confidence Bound
def upper_confidence_bound(X, X_sample, y_sample, gp, kappa=2.0):
    """
    Computes the Upper Confidence Bound (UCB) acquisition function for Bayesian Optimization.
    
    Parameters:
    X : np.ndarray
        Points where the UCB acquisition function should be evaluated, shape (n_points, n_features).
    X_sample : np.ndarray
        Sample locations already evaluated, shape (n_sample, n_features).
    y_sample : np.ndarray
        Sample values at X_sample, shape (n_sample,).
    gp : GaussianProcessRegressor
        A trained Gaussian Process model.
    kappa : float
        The exploration-exploitation trade-off parameter.
    
    Returns:
    np.ndarray
        UCB values for each point in X, shape (n_points,).
    """
    # Predict mean and standard deviation at each point X
    mu, sigma = gp.predict(X, return_std=True)
    # UCB calculation
    ucb = mu - kappa * abs(sigma)
    return ucb


